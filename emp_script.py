# app2.py — Espace Employé : Affichage automatique si matricule transmis

import streamlit as st
import pandas as pd

import streamlit as st
import json
import pandas as pd
import requests
from streamlit_lottie import st_lottie
from openpyxl import Workbook
from openpyxl import load_workbook  # Assurez-vous que cette ligne est présente
import os
import plotly.express as px

st.write("""
# Mon Espace RH - Aveni-Ré
 *Powered by Yann BEUGRE.*
""")

### LOOTIE
def load_lottiefile(filepath: str):
    with open(filepath, "r") as f:
        return json.load(f)

def load_lottieurl(url: str):
    r = requests.get(url)
    if r.status_code !=200:
        return None
    return r.json()

lottie_coding = load_lottieurl("https://assets4.lottiefiles.com/packages/lf20_x62chJ.json")  # Exemple d’animation
st_lottie(
    lottie_coding,
    speed=1, 
    reverse=False,
    loop=True,
    quality="high",
    height=None,
    width=None,
    key=None,

)

st.sidebar.title("Employé")


### DEMANDE DE CONGES
# Fonction pour ajouter une ligne de données à une feuille existante ou créer une nouvelle feuille
def add_data_to_excel(file_name, sheet_name, data_dict):
    # Vérifier si le fichier Excel existe
    if os.path.exists(file_name):
        # Charger le fichier Excel existant
        wb = load_workbook(file_name)
        
        # Vérifier si la feuille existe déjà
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            # Si la feuille n'existe pas, créer une nouvelle feuille
            sheet = wb.create_sheet(sheet_name)
            # Ajouter les en-têtes de colonne à la feuille
            sheet.append(list(data_dict.keys()))
        
        # Ajouter les nouvelles données sous forme de liste dans la feuille
        sheet.append(list(data_dict.values()))
        
        # Sauvegarder les modifications dans le fichier
        wb.save(file_name)
    else:
        # Si le fichier n'existe pas, créer un nouveau fichier Excel
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        # Ajouter les en-têtes de colonne
        sheet.append(list(data_dict.keys()))
        # Ajouter les données
        sheet.append(list(data_dict.values()))
        wb.save(file_name)

# Vérifier si "conges_clicked" existe déjà dans session_state, sinon initialisez-le
if 'conges_clicked' not in st.session_state:
    st.session_state.conges_clicked = False

# Ajouter un bouton dans la sidebar pour "Demande de congés"
conges = st.sidebar.button("Demande de congés")

# Si l'utilisateur clique sur le bouton, on change l'état
if conges:
    st.session_state.conges_clicked = True

# Afficher les champs supplémentaires si l'utilisateur a cliqué sur "Demande de congés"
if st.session_state.conges_clicked:
    # Afficher un message d'instruction
    st.sidebar.text("Veuillez remplir les informations suivantes :")
    
    # Afficher une zone de texte pour la raison du congé
    description = st.sidebar.text_area("Raison du congé", "Entrez la raison ici...")
    
    # Afficher un champ de saisie pour la date de début
    start_date = st.sidebar.date_input("Date de début")
    
    # Afficher un champ de saisie pour la date de fin
    end_date = st.sidebar.date_input("Date de fin")

    # Bouton pour soumettre la demande
    submit_conges = st.sidebar.button("Soumettre la demande")

    if submit_conges:
        # Vérifier que la date de fin est postérieure à la date de début
        if end_date < start_date:
            st.sidebar.error("La date de fin doit être postérieure à la date de début.")
        else:
            st.sidebar.success("Votre demande de congés a été envoyée avec succès!")
            # Afficher les données saisies
            st.sidebar.write(f"Raison : {description}")
            st.sidebar.write(f"Date de début : {start_date}")
            st.sidebar.write(f"Date de fin : {end_date}")

            # Enregistrer les données dans un fichier Excel
            data_conges = {
                "raison": description,
                "date_debut": start_date,
                "date_fin": end_date
            }
            file_name = "employe.xlsx"  # Ce fichier sera enregistré sur le serveur local où Streamlit est exécuté
            sheet_name = "conges"
            add_data_to_excel(file_name, sheet_name, data_conges)

            # Informer l'utilisateur que le fichier est sauvegardé
            st.sidebar.write(f"Le fichier {file_name} a été sauvegardé sur le serveur.")

            # Réinitialiser l'état après soumission pour masquer les champs de saisie
            st.session_state.conges_clicked = False



###                        FICHE DE PAIE
# Fonction pour ajouter une ligne de données à une feuille existante ou créer une nouvelle feuille
def add_data_to_excel(file_name, sheet_name, data_dict):
    # Vérifier si le fichier Excel existe
    if os.path.exists(file_name):
        # Charger le fichier Excel existant
        wb = load_workbook(file_name)
        
        # Vérifier si la feuille existe déjà
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            # Si la feuille n'existe pas, créer une nouvelle feuille
            sheet = wb.create_sheet(sheet_name)
            # Ajouter les en-têtes de colonne à la feuille
            sheet.append(list(data_dict.keys()))
        
        # Ajouter les nouvelles données sous forme de liste dans la feuille
        sheet.append(list(data_dict.values()))
        
        # Sauvegarder les modifications dans le fichier
        wb.save(file_name)
    else:
        # Si le fichier n'existe pas, créer un nouveau fichier Excel
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        # Ajouter les en-têtes de colonne
        sheet.append(list(data_dict.keys()))
        # Ajouter les données
        sheet.append(list(data_dict.values()))
        wb.save(file_name)

# Vérifier si "fiche_paie_clicked" existe déjà dans session_state, sinon initialisez-le
if 'fiche_paie_clicked' not in st.session_state:
    st.session_state.fiche_paie_clicked = False

# Ajouter un bouton dans la sidebar pour "Demande de fiche de paie"
fiche_paie = st.sidebar.button("Demande de fiche de paie")

# Si l'utilisateur clique sur le bouton, on change l'état
if fiche_paie:
    st.session_state.fiche_paie_clicked = True

# Afficher les champs supplémentaires si l'utilisateur a cliqué sur "Demande de fiche de paie"
if st.session_state.fiche_paie_clicked:
    # Afficher un message d'instruction
    st.sidebar.text("Veuillez remplir les informations suivantes :")
    
    # Afficher un champ de saisie pour la période de début de la fiche de paie
    date_debut = st.sidebar.date_input("Date de début de la période")
    
    # Afficher un champ de saisie pour la période de fin de la fiche de paie
    date_fin = st.sidebar.date_input("Date de fin de la période")

    # Bouton pour soumettre la demande
    submit_fiche_paie = st.sidebar.button("Soumettre la demande")

    if submit_fiche_paie:
        # Vérifier que la date de fin est postérieure à la date de début
        if date_fin < date_debut:
            st.sidebar.error("La date de fin doit être postérieure à la date de début.")
        else:
            st.sidebar.success("Votre demande de fiche de paie a été envoyée avec succès!")
            # Afficher la période demandée
            st.sidebar.write(f"Période demandée : Du {date_debut.strftime('%d %B %Y')} au {date_fin.strftime('%d %B %Y')}")

            # Enregistrer les données dans un fichier Excel
            data_fiche_paie = {
                "date_debut": date_debut,
                "date_fin": date_fin
            }
            file_name = "employe.xlsx"  # Ce fichier sera enregistré sur le serveur local où Streamlit est exécuté
            sheet_name = "fiche_paie"
            add_data_to_excel(file_name, sheet_name, data_fiche_paie)

            # Informer l'utilisateur que le fichier est sauvegardé
            st.sidebar.write(f"Le fichier {file_name} a été sauvegardé sur le serveur.")

            # Réinitialiser l'état après soumission pour masquer les champs de saisie
            st.session_state.fiche_paie_clicked = False

###                             FRAIS DE MISSION
# Fonction pour ajouter une ligne de données à une feuille existante ou créer une nouvelle feuille
def add_data_to_excel(file_name, sheet_name, data_dict):
    # Vérifier si le fichier Excel existe
    if os.path.exists(file_name):
        # Charger le fichier Excel existant
        wb = load_workbook(file_name)
        
        # Vérifier si la feuille existe déjà
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            # Si la feuille n'existe pas, créer une nouvelle feuille
            sheet = wb.create_sheet(sheet_name)
            # Ajouter les en-têtes de colonne à la feuille
            sheet.append(list(data_dict.keys()))
        
        # Ajouter les nouvelles données sous forme de liste dans la feuille
        sheet.append(list(data_dict.values()))
        
        # Sauvegarder les modifications dans le fichier
        wb.save(file_name)
    else:
        # Si le fichier n'existe pas, créer un nouveau fichier Excel
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        # Ajouter les en-têtes de colonne
        sheet.append(list(data_dict.keys()))
        # Ajouter les données
        sheet.append(list(data_dict.values()))
        wb.save(file_name)

# Vérifier si "frais_mission_clicked" existe déjà dans session_state, sinon initialisez-le
if 'frais_mission_clicked' not in st.session_state:
    st.session_state.frais_mission_clicked = False

# Ajouter un bouton dans la sidebar pour "Remboursement de frais de mission"
frais_mission = st.sidebar.button("Remboursement de frais de mission")

# Si l'utilisateur clique sur le bouton, on change l'état
if frais_mission:
    st.session_state.frais_mission_clicked = True

# Afficher les champs supplémentaires si l'utilisateur a cliqué sur "Remboursement de frais de mission"
if st.session_state.frais_mission_clicked:
    # Afficher un message d'instruction
    st.sidebar.text("Veuillez remplir les informations suivantes :")
    
    # Afficher un champ pour le nom de la mission / pays
    mission_pays = st.sidebar.text_input("Nom de la mission / Pays")
    
    # Afficher un champ pour la période de la mission (date de début)
    start_date_mission = st.sidebar.date_input("Date de début de la mission")
    
    # Afficher un champ pour la date de fin de la mission
    end_date_mission = st.sidebar.date_input("Date de fin de la mission")
    
    # Afficher un champ pour le montant des frais de mission
    montant_frais = st.sidebar.number_input("Montant des frais (en XOF)", min_value=0.0, format="%.2f")
    
    # Afficher un champ pour télécharger la pièce justificative
    justificatif = st.sidebar.file_uploader("Téléchargez la pièce justificative", type=["pdf", "jpg", "png"])

    # Bouton pour soumettre la demande
    submit_frais_mission = st.sidebar.button("Soumettre la demande")

    if submit_frais_mission:
        # Vérifier que la date de fin est postérieure à la date de début
        if end_date_mission < start_date_mission:
            st.sidebar.error("La date de fin de la mission doit être postérieure à la date de début.")
        else:
            st.sidebar.success("Votre demande de remboursement de frais de mission a été envoyée avec succès!")
            # Afficher les informations saisies
            st.sidebar.write(f"Nom de la mission / Pays : {mission_pays}")
            st.sidebar.write(f"Période : Du {start_date_mission.strftime('%d %B %Y')} au {end_date_mission.strftime('%d %B %Y')}")
            st.sidebar.write(f"Montant des frais : {montant_frais} XOF")
            
            if justificatif is not None:
                st.sidebar.write(f"Pièce justificative téléchargée : {justificatif.name}")

            # Enregistrer les données dans un fichier Excel
            data_frais_mission = {
                "mission_pays": mission_pays,
                "date_debut_mission": start_date_mission,
                "date_fin_mission": end_date_mission,
                "montant_frais": montant_frais,
                "justificatif": justificatif.name if justificatif else "Non fourni"
            }
            file_name = "employe.xlsx"  # Ce fichier sera enregistré sur le serveur local où Streamlit est exécuté
            sheet_name = "frais_mission"
            add_data_to_excel(file_name, sheet_name, data_frais_mission)

            # Informer l'utilisateur que le fichier est sauvegardé
            st.sidebar.write(f"Le fichier {file_name} a été sauvegardé sur le serveur.")

            # Réinitialiser l'état après soumission pour masquer les champs de saisie
            st.session_state.frais_mission_clicked = False



###                         AVANCE SUR SALAIRE
# Fonction pour ajouter une ligne de données à une feuille existante ou créer une nouvelle feuille
def add_data_to_excel(file_name, sheet_name, data_dict):
    # Vérifier si le fichier Excel existe
    if os.path.exists(file_name):
        # Charger le fichier Excel existant
        wb = load_workbook(file_name)
        
        # Vérifier si la feuille existe déjà
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            # Si la feuille n'existe pas, créer une nouvelle feuille
            sheet = wb.create_sheet(sheet_name)
            # Ajouter les en-têtes de colonne à la feuille
            sheet.append(list(data_dict.keys()))
        
        # Ajouter les nouvelles données sous forme de liste dans la feuille
        sheet.append(list(data_dict.values()))
        
        # Sauvegarder les modifications dans le fichier
        wb.save(file_name)
    else:
        # Si le fichier n'existe pas, créer un nouveau fichier Excel
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        # Ajouter les en-têtes de colonne
        sheet.append(list(data_dict.keys()))
        # Ajouter les données
        sheet.append(list(data_dict.values()))
        wb.save(file_name)

# Vérifier si "avance_salaire_clicked" existe déjà dans session_state, sinon initialisez-le
if 'avance_salaire_clicked' not in st.session_state:
    st.session_state.avance_salaire_clicked = False

# Ajouter un bouton dans la sidebar pour "Demande d'avance sur salaire"
avance_salaire = st.sidebar.button("Demande avance sur salaire")

# Si l'utilisateur clique sur le bouton, on change l'état
if avance_salaire:
    st.session_state.avance_salaire_clicked = True

# Afficher les champs supplémentaires si l'utilisateur a cliqué sur "Demande avance sur salaire"
if st.session_state.avance_salaire_clicked:
    # Afficher un message d'instruction
    st.sidebar.text("Veuillez remplir les informations suivantes :")
    
    # Afficher un champ pour le motif de l'avance sur salaire
    motif_avance = st.sidebar.text_area("Motif de la demande d'avance", "Entrez le motif ici...")
    
    # Afficher un champ pour le montant de l'avance
    montant_avance = st.sidebar.number_input("Montant de l'avance (en XOF)", min_value=0.0, format="%.2f")
    
    # Bouton pour soumettre la demande
    submit_avance_salaire = st.sidebar.button("Soumettre la demande")

    if submit_avance_salaire:
        st.sidebar.success("Votre demande d'avance sur salaire a été envoyée avec succès!")
        # Afficher les informations saisies
        st.sidebar.write(f"Motif de la demande : {motif_avance}")
        st.sidebar.write(f"Montant de l'avance : {montant_avance} XOF")
        
        # Enregistrer les données dans un fichier Excel
        data_avance_salaire = {
            "motif_avance": motif_avance,
            "montant_avance": montant_avance
        }
        file_name = "employe.xlsx"  # Ce fichier sera enregistré sur le serveur local où Streamlit est exécuté
        sheet_name = "avance_salaire"
        add_data_to_excel(file_name, sheet_name, data_avance_salaire)

        # Informer l'utilisateur que le fichier est sauvegardé
        st.sidebar.write(f"Le fichier {file_name} a été sauvegardé sur le serveur.")

        # Réinitialiser l'état après soumission pour masquer les champs de saisie
        st.session_state.avance_salaire_clicked = False

@st.cache_data
def charger_donnees():
    employes = pd.read_excel("C:\\Users\\beugre\\Desktop\\base_employes.xlsx", sheet_name="Employes")
    absences = pd.read_excel("C:\\Users\\beugre\\Desktop\\base_employes.xlsx", sheet_name="Absences")
    formations = pd.read_excel("C:\\Users\\beugre\\Desktop\\base_employes.xlsx", sheet_name="Formations")
    return employes, absences, formations

employes, absences, formations = charger_donnees()

st.title("Espace Personnel - AVENI-Ré")

query_params = st.experimental_get_query_params()
matricule = query_params.get('matricule', [None])[0]

if matricule:
    if matricule in employes['Matricule'].astype(str).values:
        employe = employes[employes['Matricule'].astype(str) == matricule].iloc[0]

        st.header(f"Bienvenue {employe['Nom']} !")
        st.subheader(f"Poste : {employe['Fonction']}")

        st.markdown("""
        **Informations personnelles :**
        - Matricule : {0}
        - Département : {1}
        - Date d'embauche : {2}
        - Contrat : {3}
        - Email : {4}
        - Responsable hiérarchique : {5}
        """.format(
            employe['Matricule'], employe['Département'], employe['Date_Embauche'],
            employe['Contrat'], employe['Email'], employe['Responsable']
        ))

        st.divider()
        st.subheader("Vos statistiques RH")
        st.metric("Solde de congés restants", f"{employe['Solde_Conge']} jours")

        st.divider()
        st.subheader("Historique des absences")
        abs_emp = absences[absences['Matricule'].astype(str) == matricule]
        st.dataframe(abs_emp[['Date', 'Type', 'Motif']])

        st.divider()
        st.subheader("Formations suivies")
        form_emp = formations[formations['Matricule'].astype(str) == matricule]
        if form_emp.empty:
            st.info("Aucune formation suivie.")
        else:
            fig3 = px.scatter(
                form_emp,
                x='Date',
                y='Intitulé Formation',
                color='Organisme',
                size_max=20,
                title="Timeline des Formations Suivies"
            )
            st.plotly_chart(fig3, use_container_width=True)

        st.dataframe(form_emp[['Date', 'Intitulé Formation', 'Organisme']])
    else:
        st.error("Matricule non reconnu.")
else:
    st.warning("Aucun matricule détecté dans l'URL. Merci de passer par la page de connexion.")

with st.expander("See more !"):
    st.write("Consulter notre charte interne :")
    st.write("Visiter notre site web :","https://www.aveni-re.com/")
